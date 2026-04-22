"""Excel output generation with percentage tables, heatmaps, and audit sheets."""

import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

from . import config

logger = logging.getLogger(__name__)

# Styling constants
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Threshold column headers for display
THRESHOLD_HEADERS = [f"RRP < ${t}" if t == 0 else f"RRP < -${abs(t)}" for t in config.THRESHOLDS]

# Column suffixes matching analyse.py output
THRESHOLD_SUFFIXES = ["0", "neg10", "neg20", "neg30", "neg40", "neg50", "neg60", "neg70", "neg80"]


def generate_all_workbooks(summary: pd.DataFrame, output_dir: str):
    """Generate one .xlsx workbook per region from the summary DataFrame."""
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    for region in config.REGIONS:
        region_data = summary[summary["REGIONID"] == region].copy()
        if region_data.empty:
            logger.warning(f"No data for {region}, skipping workbook")
            continue

        region_data = region_data.sort_values("YEAR_MONTH").reset_index(drop=True)
        friendly_name = config.REGION_NAMES[region]
        filepath = output_path / f"{friendly_name}_negative_prices.xlsx"

        _write_region_workbook(region_data, friendly_name, filepath)
        logger.info(f"Written {filepath}")

    generate_all_states_workbook(summary, output_dir)


def generate_all_states_workbook(summary: pd.DataFrame, output_dir: str):
    """Generate a single workbook with all regions as separate sheets (Percentages view)."""
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    filepath = output_path / "All_States_negative_prices.xlsx"

    wb = Workbook()

    for region in config.REGIONS:
        region_data = summary[summary["REGIONID"] == region].copy()
        if region_data.empty:
            continue
        region_data = region_data.sort_values("YEAR_MONTH").reset_index(drop=True)
        friendly_name = config.REGION_NAMES[region]
        _write_pct_sheet(wb, region_data, friendly_name, sheet_title=friendly_name)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(filepath)
    logger.info(f"Written {filepath}")


def _write_region_workbook(data: pd.DataFrame, region_name: str, filepath: Path):
    """Write a 3-sheet workbook for a single region."""
    wb = Workbook()

    # Sheet 1: Percentages
    _write_pct_sheet(wb, data, region_name)

    # Sheet 2: Heatmap
    _write_heatmap_sheet(wb, data, region_name)

    # Sheet 3: Audit / Counts
    _write_audit_sheet(wb, data, region_name)

    # Remove default empty sheet if it exists
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(filepath)


def _format_month_label(year_month: str) -> str:
    """Convert '2019-05' to 'May 2019'."""
    dt = pd.Timestamp(year_month + "-01")
    return dt.strftime("%b %Y")


def _write_pct_sheet(wb: Workbook, data: pd.DataFrame, region_name: str, sheet_title: str = "Percentages"):
    """Sheet 1: Clean percentage table."""
    ws = wb.create_sheet(title=sheet_title)

    # Header row
    headers = ["Month"] + THRESHOLD_HEADERS
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, (_, row) in enumerate(data.iterrows(), 2):
        month_cell = ws.cell(row=row_idx, column=1, value=_format_month_label(row["YEAR_MONTH"]))
        month_cell.font = Font(size=11)
        month_cell.border = THIN_BORDER

        for col_idx, suffix in enumerate(THRESHOLD_SUFFIXES, 2):
            val = row[f"pct_below_{suffix}"]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

    # Column widths
    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13

    # Freeze top row
    ws.freeze_panes = "A2"


def _write_heatmap_sheet(wb: Workbook, data: pd.DataFrame, region_name: str):
    """Sheet 2: Same data as percentages but with conditional colour formatting."""
    ws = wb.create_sheet(title="Heatmap")

    # Header row
    headers = ["Month"] + THRESHOLD_HEADERS
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Data rows
    num_rows = len(data)
    for row_idx, (_, row) in enumerate(data.iterrows(), 2):
        month_cell = ws.cell(row=row_idx, column=1, value=_format_month_label(row["YEAR_MONTH"]))
        month_cell.font = Font(size=11)
        month_cell.border = THIN_BORDER

        for col_idx, suffix in enumerate(THRESHOLD_SUFFIXES, 2):
            val = row[f"pct_below_{suffix}"]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

    # Apply colour scale per column (green=low → yellow=mid → red=high)
    if num_rows > 0:
        for col_idx in range(2, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            cell_range = f"{col_letter}2:{col_letter}{num_rows + 1}"
            ws.conditional_formatting.add(
                cell_range,
                ColorScaleRule(
                    start_type="min", start_color="63BE7B",  # green
                    mid_type="percentile", mid_value=50, mid_color="FFEB84",  # yellow
                    end_type="max", end_color="F8696B",  # red
                ),
            )

    # Column widths
    ws.column_dimensions["A"].width = 14
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13

    ws.freeze_panes = "A2"


def _write_audit_sheet(wb: Workbook, data: pd.DataFrame, region_name: str):
    """Sheet 3: Raw counts and total intervals for verification."""
    ws = wb.create_sheet(title="Audit")

    # Header row
    count_headers = [f"Count {h}" for h in THRESHOLD_HEADERS]
    headers = ["Month", "Total Daylight Intervals"] + count_headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, (_, row) in enumerate(data.iterrows(), 2):
        ws.cell(row=row_idx, column=1, value=_format_month_label(row["YEAR_MONTH"])).border = THIN_BORDER
        ws.cell(row=row_idx, column=2, value=int(row["total_daylight_intervals"])).border = THIN_BORDER

        for col_idx, suffix in enumerate(THRESHOLD_SUFFIXES, 3):
            val = int(row[f"count_below_{suffix}"])
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    for col_idx in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    ws.freeze_panes = "A2"
